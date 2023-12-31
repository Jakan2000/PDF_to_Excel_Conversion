from datetime import datetime

import openpyxl

from CommonClass import Excel


def deleteNoneRows(wb, start, end, refColumn):
    sheet = wb.active
    for x in range(end, start, -1):
        a_cell = f"A{x}"
        if sheet[a_cell].value is None:
            end -= 1
            sheet.delete_rows(x)
    return wb


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%m-%Y").date()
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(end, start, -1):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(len(dataToMerge) - 1, 0, -1):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(len(dataToMerge) - 1, 0, -1):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def axis1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 7
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def axis1_main(wb):
    sheet = wb.active
    if axis1_validation(wb):
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response
    else:
        startText = "Particulars"
        endText = "CLOSING BALANCE"
        startEndDefColumn = "C"
        delRefText1 = "OPENING BALANCE"
        delRefText2 = "TRANSACTION TOTAL"
        delRefText3 = "CLOSING BALANCE"
        deleteFlagRefColumn = "C"
        stringAlignColumn1 = "C"
        stringAlignColumn2 = "G"
        dateConversionColumn = "A"
        refHeaderText1 = "Tran Date"
        refHeaderText2 = "Chq No"
        refHeaderText3 = "Particulars"
        refHeaderText4 = "Debit"
        refHeaderText5 = "Credit"
        refHeaderText6 = "Balance"
        headerText1 = "Transaction_Date"
        headerText2 = "ChequeNo_RefNo"
        headerText3 = "Narration"
        headerText4 = "Withdrawal"
        headerText5 = "Deposit"
        headerText6 = "Balance"
        deleteColumnRefText = "Init.Br"
        columns = ["Sl.No.", "Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal",
                   "Balance"]
        negativeValueColumnRefText1 = "Withdrawal"
        headerTextToEmptyCellToNone1 = "Value_Date"
        headerTextToEmptyCellToNone2 = "ChequeNo_RefNo"
        headerTextToEmptyCellToNone3 = "Withdrawal"
        headerTextToEmptyCellToNone4 = "Deposit"
        refColumnToMerg = "A"
        mergingColumn = "C"
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        delFooter = deleteFooter(wb, end)
        headerDeleted = deleteHeader(delFooter, start - 1)
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndDefColumn)
        removed1 = Excel.remove_row(headerDeleted, start, end, delRefText1, deleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(removed1, startText, endText, startEndDefColumn)
        removed2 = Excel.remove_row(removed1, start, end, delRefText2, deleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(removed2, startText, endText, startEndDefColumn)
        removed3 = Excel.remove_row(removed1, start, end, delRefText3, deleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(removed2, startText, endText, startEndDefColumn)
        alignedC = Excel.string_align(removed3, start, end + 1, stringAlignColumn1)  # end+1 to Include Last Row
        alignedG = Excel.string_align(alignedC, start, end + 1, stringAlignColumn2)  # end+1 to Include Last Row
        mergedColumnC = mergingRows(alignedG, start, end, refColumnToMerg, mergingColumn)
        noneRowsDeleted = deleteNoneRows(mergedColumnC, start, end, refColumnToMerg)
        start, end = Excel.get_start_end_row_index(removed2, startText, endText, startEndDefColumn)
        convertedDateA = dateConvertion(noneRowsDeleted, start + 1, end + 1, dateConversionColumn)  # start+1 to Sip Header, end+1 to Include Last Row
        lastCol = 65 + sheet.max_column  # 65 => ASCII value "A"
        trandate = Excel.alter_header_name(convertedDateA, refHeaderText1, headerText1, lastCol)
        chqno = Excel.alter_header_name(trandate, refHeaderText2, headerText2, lastCol)
        naration = Excel.alter_header_name(chqno, refHeaderText3, headerText3, lastCol)
        debit = Excel.alter_header_name(naration, refHeaderText4, headerText4, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)
        balance = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)
        deletedColumnG = Excel.delete_column(balance, deleteColumnRefText)
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A"
        slnoCreated = Excel.create_slno_column(deletedColumnG, start, end + 1, chr(columnToCreateSlNo))
        columnFinalised = Excel.finalise_column(slnoCreated, columns)
        negativeValueChecked = Excel.check_neagativeValue_by_column(columnFinalised, negativeValueColumnRefText1)
        valueDateConverted = Excel.empty_cell_to_none(negativeValueChecked, start, end + 1, headerTextToEmptyCellToNone1)
        chqnoConverted = Excel.empty_cell_to_none(valueDateConverted, start, end + 1, headerTextToEmptyCellToNone2)
        withdrawalConverted = Excel.empty_cell_to_none(chqnoConverted, start, end + 1, headerTextToEmptyCellToNone3)
        depositConverted = Excel.empty_cell_to_none(withdrawalConverted, start, end + 1, headerTextToEmptyCellToNone4)
        createdTransTypeColumn = Excel.transaction_type_column(depositConverted)
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/1.Axis_-_8874-PW_-_GNAN842166790_unlocked__19-09-2023-14-05-39.xlsx"
    path = "C:/Users/Admin/Downloads/1.SVTTransports-AXIS1437__23-11-2023-17-46-06.xlsx"
    wb = openpyxl.load_workbook(path)
    result = axis1_main(wb)
    # result.save('C:/Users/Admin/Desktop/FinalOutput/AXIS1output.xlsx')
    result.save('C:/Users/Admin/Desktop/AXIS1output.xlsx')
